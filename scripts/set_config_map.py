#!/usr/bin/env python
# -*- coding:utf-8 -*-
import MySQLdb
import sys
import json
import re
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf8')

# 定义excel文件的路径
EXCEL_PATH = '/Users/pavoooo/Desktop/config.xlsx'
# 定义数据库的连接信息
db_ip = '127.0.0.1'
db_user = 'root'
db_pwd = ''
db_name = 'metabase_online_time'
db_port = 3306

db = None

db_name_map = {
    'hive': 'BI-hive',
    'logdata': 'DI-report'
}

# 比较所有的数据是不是全部相等
def all_number_is_equal(*args):
    if len(args) == 0 or len(args) == 1:
        return True
    else:
        return max(args) == min(args)

# 创建数据库操作
def connect_db_get_data():
    global db
    db = MySQLdb.connect(
        host = db_ip,
        user = db_user,
        passwd = db_pwd,
        db = db_name,
        port = db_port
    )
    # 获取油标
    cursor = db.cursor()
    format_datas = {}
    # 获取所有的数据库和表的信息
    cursor.execute('SELECT db.id, db.name, t.id, t.name FROM metabase_database AS db LEFT JOIN metabase_table AS t ON db.id = t.db_id')
    all_table_fields = cursor.fetchall()
    # 处理数据
    for fields in all_table_fields:
        (database_id, database_name, table_id, table_name) = fields
        padding_values = {
            'id': database_id,
            table_name: table_id
        }
        if len(format_datas.get(database_name, [])) == 0:
            format_datas[database_name] = [padding_values]
        else:
            format_datas[database_name].append(padding_values)
    return format_datas

# 读取excel文件
def load_excel_data():
    wb = load_workbook(EXCEL_PATH)
    sheetnames = wb.sheetnames
    formated_excel_datas = []
    if len(sheetnames) == 0:
        exit('打开的excel存在问题')
    else:
        sheet = wb.get_sheet_by_name(sheetnames[0])
        table_name_len = len(sheet['C'])
        table_data_len = len(sheet['D'])
        db_name_len = len(sheet['E'])
        table_name_title = len(sheet['B'])
        if not all_number_is_equal(table_name_len, table_data_len, db_name_len, table_name_title):
            exit('excel中的数据存在问题') 
        else:
            for i in range(table_name_len):
                formated_excel_datas.append({
                    'table_name': sheet['C'][i].value,
                    'table_datas': sheet['D'][i].value,
                    'db_name': sheet['E'][i].value,
                    'table_title': sheet['B'][i].value
                })
            return formated_excel_datas

def format_table_data():
    excel_datas = load_excel_data()
    try:
        for data in excel_datas:
            # fields = json.loads(data['table_datas']).get('data_table', [])
            fields = re.findall(r'"data_table":(.+)\}$', data['table_datas'])
            fields = [] if len(fields) == 0 else json.loads(fields[0])
            if isinstance(fields, dict):
                to_list = []
                keys = fields.keys()
                for key in keys:
                    to_list = to_list + fields[key]
                fields = to_list
            formated_fields = {}
            for field in fields:
                chinese_field = field[0]
                english_field = field[1]
                if english_field:
                    formated_fields[english_field] = chinese_field
            data['table_datas'] = formated_fields
    except Exception as e:
        print data
        exit('数据有误')
    return excel_datas

def mixin_data():
    db_datas = connect_db_get_data()
    excel_datas = format_table_data()
    finall_datas = []
    for data in excel_datas:
        db_name = db_name_map.get(data['db_name'], data['db_name'])
        table_name = data['table_name']
        table_datas = data['table_datas']
        table_title = data['table_title']
        has_table = False
        if len(db_datas.get(db_name, [])) > 0:
            # print '^'*10, table_name, '^'*10 
            db_tables = db_datas.get(db_name)
            for table in db_tables:
                if table.has_key(table_name):
                    has_table = True
                    finall_datas.append({
                        'db_id': table['id'],
                        'table_id': table[table_name],
                        'fields_map': table_datas,
                        'table_name': table_name,
                        'table_title': table_title
                    })
                    break
        if not has_table:
            print table_name, '*'*10, table_title
        # else:
        #     print '='*10, table_name, '='*10
    return finall_datas
def insert_data():
    global db
    if db is None:
        db = MySQLdb.connect(
            host = db_ip,
            user = db_user,
            passwd = db_pwd,
            db = db_name,
            port = db_port
        )
    computed_datas = mixin_data()
    # exit()
    cursor = db.cursor()
    try:
        for data in computed_datas:
            cursor.execute("INSERT INTO field_map (db_id, table_id, table_name, table_title, fields_map) values (%s, %s, \'%s\', \'%s\', \'%s\')" % (data['db_id'], data['table_id'], data['table_name'], data['table_title'], json.dumps(data['fields_map'], ensure_ascii = False)))
        print '插入成功'
        db.commit()
    except Exception as e:
        print("INSERT INTO fields_map (db_id, table_id, table_name, table_title, fields_map) values (%s, %s, \'%s\' ,\'%s\', \'%s\')" % (data['db_id'], data['table_id'], data['table_name'], data['table_title'], json.dumps(data['fields_map'])))
        db.rollback()
        print e
        exit('数据插入出错')
    
if __name__ == '__main__':
    def format_output(json_data):
        return json.dumps(json_data, ensure_ascii = False, indent = 2)
    insert_data()
    if db is not None:
        db.close()